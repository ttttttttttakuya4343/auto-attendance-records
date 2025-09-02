#!/usr/bin/env python3
"""
Simple アプリケーションのテストコード
"""

import unittest
import sys
import os
import subprocess
from unittest.mock import patch


class TestSimpleApp(unittest.TestCase):

    def setUp(self):
        """テスト前の準備"""
        self.test_csv_path = "tests/test_data/test_attendance.csv"
        self.test_template_path = (
            "tests/test_data/Quco_テスト_2025年8月度勤務表・立替経費精算書.xlsx"
        )

    @patch("builtins.input")
    def test_simple_app_execution(self, mock_input):
        """Simple アプリケーション実行テスト"""
        # 入力をモック
        mock_input.side_effect = [
            "テストユーザー",  # 作業者名
            self.test_csv_path,  # CSVファイルパス
            self.test_template_path,  # テンプレートファイルパス
        ]

        # 出力ディレクトリを作成
        os.makedirs("data/output", exist_ok=True)

        try:
            # Simple アプリケーションを実行
            cmd = [sys.executable, "apps/simple_app.py"]
            result = subprocess.run(
                cmd,
                input=f"テストユーザー\n{self.test_csv_path}\n{self.test_template_path}\n",
                capture_output=True,
                text=True,
                cwd=".",
            )

            # 実行結果の確認
            self.assertEqual(result.returncode, 0)

            # 出力ファイルの存在確認
            expected_output = (
                "data/output/Quco_テストユーザー_2025年8月度勤務表・立替経費精算書.xlsx"
            )
            self.assertTrue(os.path.exists(expected_output))

            # 出力ファイルの内容確認
            from openpyxl import load_workbook

            wb = load_workbook(expected_output)
            ws = wb["業務報告書（現場業務用）"]
            self.assertEqual(ws["C7"].value, "テストユーザー")

        finally:
            # テスト後のクリーンアップ
            output_files = [
                "data/output/Quco_テストユーザー_2025年8月度勤務表・立替経費精算書.xlsx"
            ]
            for file_path in output_files:
                if os.path.exists(file_path):
                    os.unlink(file_path)


if __name__ == "__main__":
    unittest.main()
