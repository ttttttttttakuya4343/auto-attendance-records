#!/usr/bin/env python3
"""
AttendanceProcessorのテストコード
"""

import unittest
import sys
import os
import tempfile
import re
from datetime import time

from openpyxl import load_workbook

sys.path.append("src")
from attendance_processor import AttendanceProcessor  # noqa: E402


class TestAttendanceProcessor(unittest.TestCase):

    def setUp(self):
        """テスト前の準備"""
        self.processor = AttendanceProcessor()
        self.test_csv_path = "tests/test_data/test_attendance.csv"
        self.test_template_path = (
            "tests/test_data/Quco_テスト_2025年8月度勤務表・立替経費精算書.xlsx"
        )

    def test_load_csv(self):
        """CSVファイル読み込みテスト"""
        result = self.processor.load_csv(self.test_csv_path)
        self.assertTrue(result)
        self.assertEqual(len(self.processor.csv_data), 3)
        self.assertEqual(self.processor.csv_data[0]["出勤時刻"], "09:00")

    def test_load_excel_template(self):
        """Excelテンプレート読み込みテスト"""
        result = self.processor.load_excel_template(self.test_template_path)
        self.assertTrue(result)
        self.assertIsNotNone(self.processor.workbook)
        self.assertEqual(
            self.processor.template_filename,
            "Quco_テスト_2025年8月度勤務表・立替経費精算書.xlsx",
        )

    def test_convert_time_format(self):
        """時刻形式変換テスト"""
        # 正常な時刻
        result = self.processor._convert_time_format("09:30")
        self.assertEqual(result, time(9, 30))

        # 空文字
        result = self.processor._convert_time_format("")
        self.assertIsNone(result)

        # None
        result = self.processor._convert_time_format(None)
        self.assertIsNone(result)

    def test_get_safe_value(self):
        """安全な値取得テスト"""
        test_row = {"出勤時刻": "09:00", "退勤時刻": "", "休憩時間": None}

        # 正常な値
        result = self.processor._get_safe_value(test_row, "出勤時刻")
        self.assertEqual(result, "09:00")

        # 空文字
        result = self.processor._get_safe_value(test_row, "退勤時刻")
        self.assertIsNone(result)

        # 存在しないキー
        result = self.processor._get_safe_value(test_row, "存在しないキー")
        self.assertIsNone(result)

    def test_process_attendance(self):
        """勤怠データ処理テスト"""
        # CSVとExcelを読み込み
        self.processor.load_csv(self.test_csv_path)
        self.processor.load_excel_template(self.test_template_path)

        # 一時ファイルで出力テスト
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            output_path = tmp.name

        try:
            result = self.processor.process_attendance("田中太郎", output_path)
            self.assertTrue(result)
            self.assertTrue(os.path.exists(output_path))

            # 出力ファイルの内容確認

            wb = load_workbook(output_path)
            ws = wb["業務報告書（現場業務用）"]
            self.assertEqual(ws["C7"].value, "田中太郎")

        finally:
            if os.path.exists(output_path):
                os.unlink(output_path)

    def test_filename_generation(self):
        """ファイル名生成テスト"""
        self.processor.load_excel_template(self.test_template_path)

        # output_pathを指定しない場合のテスト

        expected_filename = re.sub(
            r"Quco_[^_]+_", "Quco_山田花子_", self.processor.template_filename
        )
        self.assertEqual(
            expected_filename, "Quco_山田花子_2025年8月度勤務表・立替経費精算書.xlsx"
        )


if __name__ == "__main__":
    unittest.main()
