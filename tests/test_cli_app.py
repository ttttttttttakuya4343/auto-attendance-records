#!/usr/bin/env python3
"""
CLI アプリケーションのテストコード
"""

import unittest
import sys
import os
import tempfile
import subprocess


class TestCliApp(unittest.TestCase):

    def setUp(self):
        """テスト前の準備"""
        self.test_csv_path = "tests/test_data/test_attendance.csv"
        self.test_template_path = (
            "tests/test_data/Quco_テスト_2025年8月度勤務表・立替経費精算書.xlsx"
        )

    def test_cli_app_execution(self):
        """CLIアプリケーション実行テスト"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            output_path = tmp.name

        try:
            # CLIアプリケーションを実行
            cmd = [
                sys.executable,
                "src/cli_app.py",
                "--csv",
                self.test_csv_path,
                "--template",
                self.test_template_path,
                "--output",
                output_path,
                "--name",
                "テストユーザー",
            ]

            result = subprocess.run(cmd, capture_output=True, text=True, cwd=".")

            # 実行結果の確認
            self.assertEqual(result.returncode, 0)
            self.assertTrue(os.path.exists(output_path))

            # 出力ファイルの内容確認
            from openpyxl import load_workbook

            wb = load_workbook(output_path)
            ws = wb["業務報告書（現場業務用）"]
            self.assertEqual(ws["C7"].value, "テストユーザー")

        finally:
            if os.path.exists(output_path):
                os.unlink(output_path)

    def test_cli_app_missing_file(self):
        """存在しないファイルでのエラーテスト"""
        cmd = [
            sys.executable,
            "src/cli_app.py",
            "--csv",
            "存在しないファイル.csv",
            "--template",
            self.test_template_path,
            "--output",
            "output.xlsx",
            "--name",
            "テストユーザー",
        ]

        result = subprocess.run(cmd, capture_output=True, text=True, cwd=".")
        self.assertNotEqual(result.returncode, 0)


if __name__ == "__main__":
    unittest.main()
