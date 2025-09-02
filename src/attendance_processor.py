import csv
from openpyxl import load_workbook
from datetime import datetime


class AttendanceProcessor:
    def __init__(self):
        self.csv_data = []
        self.workbook = None

    def load_csv(self, csv_path):
        """CSVファイルを読み込み"""
        self.csv_data = []
        with open(csv_path, "r", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            for row in reader:
                self.csv_data.append(row)
        return True

    def load_excel_template(self, excel_path):
        """Excelテンプレートを読み込み"""
        self.workbook = load_workbook(excel_path)
        # テンプレートファイル名を保存
        import os

        self.template_filename = os.path.basename(excel_path)
        return True

    def process_attendance(self, employee_name, output_path=None):
        """勤怠データをExcelに転記"""
        if not self.csv_data or self.workbook is None:
            raise ValueError("CSVまたはExcelファイルが読み込まれていません")

        if output_path is None:
            # テンプレートファイル名で名前を置換（月の部分は保持）
            import re

            output_path = re.sub(
                r"Quco_[^_]+_", f"Quco_{employee_name}_", self.template_filename
            )

        ws = self.workbook["業務報告書（現場業務用）"]
        ws["C7"] = employee_name

        for i, row in enumerate(self.csv_data):
            excel_row = 13 + i

            # 出勤時刻
            start_time = self._get_safe_value(row, "出勤時刻")
            ws[f"D{excel_row}"] = self._convert_time_format(start_time)

            # 退勤時刻
            end_time = self._get_safe_value(row, "退勤時刻")
            ws[f"E{excel_row}"] = self._convert_time_format(end_time)

            # 休憩時間
            break_time = self._get_safe_value(row, "休憩時間")
            ws[f"F{excel_row}"] = self._convert_time_format(break_time)

        self.workbook.save(output_path)
        return True

    def _get_safe_value(self, row, key):
        """安全に値を取得"""
        value = row.get(key, "")
        return value if value and value.strip() else None

    def _convert_time_format(self, time_str):
        """時刻形式を変換"""
        if not time_str:
            return None
        try:
            return datetime.strptime(time_str.strip(), "%H:%M").time()
        except ValueError:
            return None
